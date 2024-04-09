from openpyxl import load_workbook
from json import dumps
from json import dump

def to_camel_case(text):
    s = text.replace("-", " ").replace("_", " ")
    s = s.split()
    if len(text) == 0:
        return text
    return s[0] + ''.join(i.capitalize() for i in s[1:])

# Load Excel workbook
wb = load_workbook("Saboo Suzuki Due Mar24.xlsx", data_only=True)
 
# Choose a specific sheet
sheet = wb["calculator"]
 
# Find the number of rows and columns in the sheet
rows = sheet.max_row
columns = sheet.max_column
 
# List to store all rows as dictionaries
lst = []
 
# Iterate over rows and columns to extract data
for i in range(3, rows):
    row = {}
    for j in range(1,16):
        column_name = sheet.cell(row=3, column=j)
        row_data = sheet.cell(row=i+1, column=j)
        value= row_data.value
        if j == 7 or j == 12:
            value =  str(row_data.value)
        # if name is null dont make a row.
    
        row.update(
            {
                to_camel_case(column_name.value): value
            }
        )
    if row['customerName'] != None and row['DueDate'] != 'next month':
        lst.append(row)
 
# Convert extracted data into JSON format
json_data = dumps(lst)
 
# Print the JSON data
# print(json_data)


#with open('data.json', 'w') as f:
#  dump(json_data, f, ensure_ascii=False)
with open('data.json', 'w') as out_file:
    out_file.write(json_data)

